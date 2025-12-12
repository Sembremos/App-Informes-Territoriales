import io
from typing import Any, Dict, List, Optional, Tuple

import pandas as pd
import streamlit as st
from openpyxl import load_workbook

st.set_page_config(page_title="Índice Territorial — Lectura masiva de Excel", layout="wide")

CSS = """
<style>
.card {
  border: 1px solid rgba(255,255,255,0.15);
  border-radius: 14px;
  padding: 16px 16px;
  background: rgba(255,255,255,0.04);
  margin-bottom: 14px;
}
.small-muted {opacity:0.75; font-size:12px;}
.section-title {font-weight:900; font-size:16px; margin: 10px 0 6px 0;}
.kpi-wrap {display:flex; gap:10px; flex-wrap:wrap; margin-top:10px;}
.kpi {
  padding: 12px;
  border-radius: 12px;
  border: 1px solid rgba(255,255,255,0.12);
  min-width: 220px;
  background: rgba(255,255,255,0.03);
}
.kpi .label {opacity:0.85; font-size:12px;}
.kpi .value {font-weight:900; font-size:22px; margin-top:2px;}
.badge {
  display:inline-block;
  padding:6px 10px;
  border-radius:999px;
  font-weight:900;
  font-size:12px;
  color:#111;
}
hr.sep {border:none; border-top:1px solid rgba(255,255,255,0.12); margin:12px 0;}
</style>
"""
st.markdown(CSS, unsafe_allow_html=True)


# ============================================================
# Normalización / conversión
# ============================================================
def norm(x: Any) -> str:
    if x is None:
        return ""
    s = str(x).strip().lower()
    s = s.replace("\n", " ").replace("\r", " ")
    s = " ".join(s.split())
    s = (
        s.replace("á", "a").replace("é", "e").replace("í", "i")
        .replace("ó", "o").replace("ú", "u").replace("ü", "u").replace("ñ", "n")
    )
    return s

def to_pct(x: Any) -> Optional[float]:
    if x is None:
        return None
    try:
        v = float(x)
    except Exception:
        return None
    if 0 <= v <= 1.5:
        v *= 100.0
    return v

def classify_index(x: float) -> str:
    if x <= 20:
        return "Crítico (0-20)"
    if x <= 40:
        return "Bajo (20,1-40)"
    if x <= 60:
        return "Medio (40,1-60)"
    if x <= 80:
        return "Alto (60,1-80)"
    return "Muy Alto (80-100)"

def level_color(level: str) -> str:
    lv = norm(level)
    if "critico" in lv:
        return "#ff4d4d"
    if "bajo" in lv:
        return "#ffb84d"
    if "medio" in lv:
        return "#ffd84d"
    if "alto" in lv:
        return "#7bdc7b"
    return "#2ea44f"


# ============================================================
# Lectura robusta con merges
# ============================================================
def sheet_to_matrix(ws) -> List[List[Any]]:
    max_r = ws.max_row or 1
    max_c = ws.max_column or 1

    merge_map = {}
    for mr in ws.merged_cells.ranges:
        top_val = ws.cell(mr.min_row, mr.min_col).value
        for r in range(mr.min_row, mr.max_row + 1):
            for c in range(mr.min_col, mr.max_col + 1):
                merge_map[(r, c)] = top_val

    mat = []
    for r in range(1, max_r + 1):
        row = []
        for c in range(1, max_c + 1):
            v = ws.cell(r, c).value
            if v is None and (r, c) in merge_map:
                v = merge_map[(r, c)]
            row.append(v)
        mat.append(row)
    return mat


# ============================================================
# Detección de tablas
# ============================================================
def find_title_cells(mat: List[List[Any]], title_needles: List[str]) -> List[Tuple[int, int]]:
    needles = [norm(n) for n in title_needles]
    hits = []
    for r, row in enumerate(mat):
        for c, cell in enumerate(row):
            t = norm(cell)
            if not t:
                continue
            for nd in needles:
                if nd and nd in t:
                    hits.append((r, c))
                    break
    return hits

def find_pct_column_near(mat: List[List[Any]], anchor_r: int, search_rows: int = 14) -> Optional[int]:
    candidates = []
    r0 = max(0, anchor_r)
    r1 = min(len(mat), anchor_r + search_rows)
    for r in range(r0, r1):
        row = mat[r]
        for c, cell in enumerate(row):
            t = norm(cell)
            if t == "%" or "porcentaje" in t:
                candidates.append(c)
    if not candidates:
        return None
    return max(set(candidates), key=candidates.count)

def read_table_down(mat: List[List[Any]], start_r: int, pct_col: int, max_rows: int = 40) -> Dict[str, float]:
    out: Dict[str, float] = {}
    r1 = min(len(mat), start_r + max_rows)

    for r in range(start_r + 1, r1):
        row = mat[r]
        pct = to_pct(row[pct_col] if pct_col < len(row) else None)

        label = ""
        for c in range(len(row)):
            txt = norm(row[c])
            if not txt:
                continue
            if txt in ("respuesta", "total", "comunidad", "comercio", "%"):
                continue
            if "porcentaje" in txt:
                continue
            label = txt
            break

        if label and pct is not None:
            out[label] = float(pct)

    return out

def match_labels(table: Dict[str, float], needed_any: List[str]) -> bool:
    keys = list(table.keys())
    hits = 0
    for n in needed_any:
        nn = norm(n)
        if any(nn == k or nn in k for k in keys):
            hits += 1
    return hits >= max(2, len(needed_any) // 2)

def get_value_contains(table: Dict[str, float], needle: str) -> float:
    nd = norm(needle)
    for k, v in table.items():
        if nd == k or nd in k:
            return float(v)
    return 0.0


# ============================================================
# Bloques a detectar
# ============================================================
BLOCKS = {
    "PG": {
        "titles": ["se siente seguro en su comunidad", "siente seguro en su comunidad"],
        "expect": ["no", "si", "sí"],
    },
    "CA": {
        "titles": ["comparacion con el ano anterior", "comparacion con el año anterior", "comparación con el año anterior"],
        "expect": ["igual", "mas seguro", "más seguro", "menos seguro"],
    },
    "SP": {
        "titles": ["percepcion del servicio policial", "percepción del servicio policial"],
        "expect": ["excelente", "buena", "regular", "mala", "muy mala"],
    },
    "UA": {
        "titles": ["calificacion del servicio policial del ultimo ano", "calificacion del servicio policial del ultimo de ano", "calificacion del servicio policial del ultimo año"],
        "expect": ["igual", "mejor", "peor"],
    },
}

# Pesos
PG_W = {"inseguro": 0.0, "seguro": 1.0}
CA_W = {"menos_seguro": 0.0, "igual": 0.5, "mas_seguro": 1.0}
SP_W = {"excelente": 1.0, "buena": 0.75, "regular": 0.50, "mala": 0.0, "muy_mala": 0.0}
UA_W = {"peor": 0.0, "igual": 0.5, "mejor": 1.0}

def score(table_map: Dict[str, float], weights: Dict[str, float]) -> float:
    return sum(float(table_map.get(k, 0.0) or 0.0) * w for k, w in weights.items())


def extract_from_workbook(wb) -> Tuple[Optional[Dict[str, Any]], Optional[Dict[str, Any]], List[str]]:
    found = {"PG": None, "CA": None, "SP": None, "UA": None}

    for sname in wb.sheetnames:
        ws = wb[sname]
        mat = sheet_to_matrix(ws)

        for key, cfg in BLOCKS.items():
            if found[key] is not None:
                continue

            title_cells = find_title_cells(mat, cfg["titles"])
            for (tr, tc) in title_cells:
                pct_col = find_pct_column_near(mat, tr, search_rows=14)
                if pct_col is None:
                    continue
                tbl = read_table_down(mat, tr, pct_col, max_rows=40)
                if match_labels(tbl, cfg["expect"]):
                    found[key] = tbl
                    break

        if all(found[k] is not None for k in found):
            break

    errors = []
    if found["PG"] is None: errors.append("No detecté la tabla de Percepción General (No/Sí).")
    if found["CA"] is None: errors.append("No detecté la tabla de Comparación Año Anterior (Menos/Igual/Más).")
    if found["SP"] is None: errors.append("No detecté la tabla de Percepción del Servicio Policial (Excelente…Muy Mala).")
    if found["UA"] is None: errors.append("No detecté la tabla de Calificación del Último Año (Igual/Mejor/Peor).")
    if errors:
        return None, None, errors

    raw = {
        "pg_no": get_value_contains(found["PG"], "no"),
        "pg_si": get_value_contains(found["PG"], "si") or get_value_contains(found["PG"], "sí"),

        "ca_menos": get_value_contains(found["CA"], "menos seguro"),
        "ca_igual": get_value_contains(found["CA"], "igual"),
        "ca_mas": get_value_contains(found["CA"], "mas seguro") or get_value_contains(found["CA"], "más seguro"),

        "sp_excelente": get_value_contains(found["SP"], "excelente"),
        "sp_buena": get_value_contains(found["SP"], "buena"),
        "sp_regular": get_value_contains(found["SP"], "regular"),
        "sp_mala": get_value_contains(found["SP"], "mala"),
        "sp_muy_mala": get_value_contains(found["SP"], "muy mala"),

        "ua_igual": get_value_contains(found["UA"], "igual"),
        "ua_mejor": get_value_contains(found["UA"], "mejor"),
        "ua_peor": get_value_contains(found["UA"], "peor"),
    }

    pg_map = {"inseguro": raw["pg_no"], "seguro": raw["pg_si"]}
    ca_map = {"menos_seguro": raw["ca_menos"], "igual": raw["ca_igual"], "mas_seguro": raw["ca_mas"]}
    sp_map = {
        "excelente": raw["sp_excelente"], "buena": raw["sp_buena"], "regular": raw["sp_regular"],
        "mala": raw["sp_mala"], "muy_mala": raw["sp_muy_mala"]
    }
    ua_map = {"igual": raw["ua_igual"], "mejor": raw["ua_mejor"], "peor": raw["ua_peor"]}

    s_pg = score(pg_map, PG_W)
    s_ca = score(ca_map, CA_W)
    s_sp = score(sp_map, SP_W)
    s_ua = score(ua_map, UA_W)

    entorno = (s_pg + s_ca) / 2.0
    policia = (s_sp + s_ua) / 2.0
    idx = (entorno + policia) / 2.0
    level = classify_index(idx)

    res = {
        "puntaje_percepcion_general": s_pg,
        "puntaje_comparacion_anio_anterior": s_ca,
        "puntaje_servicio_policial": s_sp,
        "puntaje_ultimo_anio": s_ua,
        "percepcion_del_entorno": entorno,
        "desempeno_policia": policia,
        "indice_global": idx,
        "nivel_indice": level,
    }
    return res, raw, []


# ============================================================
# UI helpers
# ============================================================
def df_block(rows: List[Tuple[str, float]]) -> pd.DataFrame:
    d = pd.DataFrame(rows, columns=["Respuesta", "Porcentaje"])
    d["Porcentaje"] = d["Porcentaje"].map(lambda x: f"{x:.2f}%")
    return d


# ============================================================
# UI
# ============================================================
st.title("Índice Territorial — Lectura masiva de Excel")
st.caption("Sí: podés cargar múltiples Excel (hasta 80) y se calculan todos.")

files = st.file_uploader(
    "Sube hasta 80 archivos Excel (.xlsx / .xlsm)",
    type=["xlsx", "xlsm"],
    accept_multiple_files=True
)

if not files:
    st.stop()

results_rows = []
fails = []

for f in files:
    try:
        wb = load_workbook(f, data_only=True)
        res, raw, errs = extract_from_workbook(wb)
        if errs:
            fails.append({"archivo": f.name, "errores": errs})
            continue

        results_rows.append({
            "archivo": f.name,

            "pg_no_%": raw["pg_no"],
            "pg_si_%": raw["pg_si"],

            "ca_menos_seguro_%": raw["ca_menos"],
            "ca_igual_%": raw["ca_igual"],
            "ca_mas_seguro_%": raw["ca_mas"],

            "sp_excelente_%": raw["sp_excelente"],
            "sp_buena_%": raw["sp_buena"],
            "sp_regular_%": raw["sp_regular"],
            "sp_mala_%": raw["sp_mala"],
            "sp_muy_mala_%": raw["sp_muy_mala"],

            "ua_igual_%": raw["ua_igual"],
            "ua_mejor_%": raw["ua_mejor"],
            "ua_peor_%": raw["ua_peor"],

            "puntaje_percepcion_general": res["puntaje_percepcion_general"],
            "puntaje_comparacion_anio_anterior": res["puntaje_comparacion_anio_anterior"],
            "puntaje_servicio_policial": res["puntaje_servicio_policial"],
            "puntaje_ultimo_anio": res["puntaje_ultimo_anio"],

            "percepcion_del_entorno": res["percepcion_del_entorno"],
            "desempeno_policia": res["desempeno_policia"],
            "indice_global": res["indice_global"],
            "nivel_indice": res["nivel_indice"],
        })

    except Exception as e:
        fails.append({"archivo": f.name, "errores": [f"Error general leyendo archivo: {e}"]})


if results_rows:
    st.subheader("✅ Resultados")

    for r in results_rows:
        level = r["nivel_indice"]
        color = level_color(level)

        st.markdown('<div class="card">', unsafe_allow_html=True)
        st.markdown(f"<div style='font-weight:900; font-size:16px;'>📄 {r['archivo']}</div>", unsafe_allow_html=True)

        st.markdown('<div class="kpi-wrap">', unsafe_allow_html=True)
        st.markdown(f"""
            <div class="kpi"><div class="label">Percepción del entorno</div><div class="value">{r["percepcion_del_entorno"]:.2f}</div></div>
            <div class="kpi"><div class="label">Desempeño policial</div><div class="value">{r["desempeno_policia"]:.2f}</div></div>
            <div class="kpi"><div class="label">Índice Global</div><div class="value">{r["indice_global"]:.2f}</div></div>
        """, unsafe_allow_html=True)
        st.markdown("</div>", unsafe_allow_html=True)

        st.markdown(f"<span class='badge' style='background:{color};'>{level}</span>", unsafe_allow_html=True)
        st.markdown("<hr class='sep'>", unsafe_allow_html=True)

        colA, colB = st.columns(2)

        with colA:
            st.markdown("<div class='section-title'>Percepción general — Datos detectados</div>", unsafe_allow_html=True)
            st.dataframe(df_block([("No", r["pg_no_%"]), ("Sí", r["pg_si_%"])]), use_container_width=True, hide_index=True)

            st.markdown("<div class='section-title'>Comparación con el año anterior — Datos detectados</div>", unsafe_allow_html=True)
            st.dataframe(
                df_block([
                    ("Menos seguro", r["ca_menos_seguro_%"]),
                    ("Igual", r["ca_igual_%"]),
                    ("Más seguro", r["ca_mas_seguro_%"]),
                ]),
                use_container_width=True,
                hide_index=True
            )

        with colB:
            st.markdown("<div class='section-title'>Percepción del servicio policial — Datos detectados</div>", unsafe_allow_html=True)
            st.dataframe(
                df_block([
                    ("Excelente", r["sp_excelente_%"]),
                    ("Buena", r["sp_buena_%"]),
                    ("Regular", r["sp_regular_%"]),
                    ("Mala", r["sp_mala_%"]),
                    ("Muy mala", r["sp_muy_mala_%"]),
                ]),
                use_container_width=True,
                hide_index=True
            )

            st.markdown("<div class='section-title'>Servicio policial del último año — Datos detectados</div>", unsafe_allow_html=True)
            st.dataframe(
                df_block([
                    ("Igual", r["ua_igual_%"]),
                    ("Mejor servicio", r["ua_mejor_%"]),
                    ("Peor servicio", r["ua_peor_%"]),
                ]),
                use_container_width=True,
                hide_index=True
            )

        st.markdown("<hr class='sep'>", unsafe_allow_html=True)

        st.markdown("<div class='section-title'>Puntajes por bloque (0–100)</div>", unsafe_allow_html=True)
        df_scores = pd.DataFrame(
            [
                ("Percepción general (No/Sí)", r["puntaje_percepcion_general"]),
                ("Comparación con el año anterior (Menos/Igual/Más)", r["puntaje_comparacion_anio_anterior"]),
                ("Percepción del servicio policial (Excelente…Muy mala)", r["puntaje_servicio_policial"]),
                ("Servicio policial del último año (Igual/Mejor/Peor)", r["puntaje_ultimo_anio"]),
            ],
            columns=["Bloque", "Puntaje (0-100)"]
        )
        df_scores["Puntaje (0-100)"] = df_scores["Puntaje (0-100)"].map(lambda x: f"{x:.2f}")
        st.dataframe(df_scores, use_container_width=True, hide_index=True)

        st.markdown(
            "<div class='small-muted'>Fórmulas: Entorno = promedio(Percepción general, Comparación). "
            "Policía = promedio(Servicio policial, Último año). Global = promedio(Entorno, Policía).</div>",
            unsafe_allow_html=True
        )
        st.markdown("</div>", unsafe_allow_html=True)

    # ============================================================
    # Consolidado con títulos CLAROS
    # ============================================================
    st.subheader("📊 Consolidado")

    df_out = pd.DataFrame(results_rows).copy()

    rename_map = {
        "archivo": "Archivo",

        "pg_no_%": "Percepción general – No (%)",
        "pg_si_%": "Percepción general – Sí (%)",

        "ca_menos_seguro_%": "Comparación año anterior – Menos seguro (%)",
        "ca_igual_%": "Comparación año anterior – Igual (%)",
        "ca_mas_seguro_%": "Comparación año anterior – Más seguro (%)",

        "sp_excelente_%": "Servicio policial – Excelente (%)",
        "sp_buena_%": "Servicio policial – Buena (%)",
        "sp_regular_%": "Servicio policial – Regular (%)",
        "sp_mala_%": "Servicio policial – Mala (%)",
        "sp_muy_mala_%": "Servicio policial – Muy mala (%)",

        "ua_igual_%": "Último año – Igual (%)",
        "ua_mejor_%": "Último año – Mejor servicio (%)",
        "ua_peor_%": "Último año – Peor servicio (%)",

        "puntaje_percepcion_general": "Puntaje percepción general",
        "puntaje_comparacion_anio_anterior": "Puntaje comparación año anterior",
        "puntaje_servicio_policial": "Puntaje servicio policial",
        "puntaje_ultimo_anio": "Puntaje último año",

        "percepcion_del_entorno": "Percepción del entorno",
        "desempeno_policia": "Desempeño policial",
        "indice_global": "Índice global",
        "nivel_indice": "Nivel del índice",
    }

    df_show = df_out.rename(columns=rename_map).copy()

    # Formato visual (manteniendo df_out numérico para export si quisieras)
    pct_cols = [c for c in df_show.columns if c.endswith("(%)")]
    for c in pct_cols:
        df_show[c] = df_show[c].map(lambda x: f"{float(x):.2f}%")

    num_cols = [
        "Puntaje percepción general", "Puntaje comparación año anterior", "Puntaje servicio policial", "Puntaje último año",
        "Percepción del entorno", "Desempeño policial", "Índice global"
    ]
    for c in num_cols:
        if c in df_show.columns:
            df_show[c] = df_show[c].map(lambda x: f"{float(x):.3f}")

    st.dataframe(df_show.sort_values("Índice global", ascending=True), use_container_width=True)

    # Descargar: también con títulos claros
    bio = io.BytesIO()
    with pd.ExcelWriter(bio, engine="openpyxl") as writer:
        # Exporto df_show pero en versión NUMÉRICA para Excel (sin % como texto)
        df_export = df_out.rename(columns=rename_map).copy()
        df_export.to_excel(writer, index=False, sheet_name="consolidado")

    st.download_button(
        "⬇️ Descargar consolidado (Excel)",
        data=bio.getvalue(),
        file_name="consolidado_indices.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

if fails:
    st.subheader("❌ Archivos que no calzaron (detalle)")
    for item in fails:
        with st.expander(item["archivo"], expanded=True):
            for e in item["errores"]:
                st.write("•", e)
